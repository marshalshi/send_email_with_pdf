# -*- coding: utf-8 -*-
import os
import re
import shutil
from StringIO import StringIO

from PyPDF2 import PdfFileWriter, PdfFileReader
from PyPDF2.pdf import ContentStream, TextStringObject
from PyPDF2.utils import isString, b_, u_, ord_, chr_, str_, formatWarning
from openpyxl import load_workbook

from django.conf import settings
from django.core.mail import EmailMessage


def convert_page_to_text(page):
    '''
    This function will copied from PyPDF2 extractText method. 
    '''
    text = u_("")
    content = page.getContents()
    if not isinstance(content, ContentStream):
        content = ContentStream(content, page.pdf)
    # Note: we check all strings are TextStringObjects.  ByteStringObjects
    # are strings where the byte->string encoding was unknown, so adding
    # them to the text here would be gibberish.
    for operands, operator in content.operations:
        if operator == b_("Tj"):
            _text = operands[0]
            if isinstance(_text, TextStringObject):
                text += _text + ' '
        elif operator == b_("T*"):
            text += "\n"
        elif operator == b_("'"):
            text += "\n"
            _text = operands[0]
            if isinstance(_text, TextStringObject):
                text += operands[0] + ' '
        elif operator == b_('"'):
            _text = operands[2]
            if isinstance(_text, TextStringObject):
                text += "\n"
                text += _text + ' '
        elif operator == b_("TJ"):
            for i in operands[0]:
                if isinstance(i, TextStringObject):
                    text += i + ' '
            text += "\n"
            
    return text

def split_pdf(f_path, split_key_word):
    '''
    This function will split pdf to several sub-pdfs according key words
    which mark down to be split key.

    Function will remove all PDF files in settings.PDF_FOLDER in order
    to clean all historical useless files.
    
    Args:
      f_path: target pdf path.
    '''
    pdf_folder = settings.PDF_FOLDER

    # Remove all existing PDFs in pdf_folder
    for f_name in os.listdir(pdf_folder):
        existing_file = os.path.join(pdf_folder, f_name)
        f_extension = os.path.splitext(f_name)[0]
        if os.path.isfile(existing_file) and f_extension == '.pdf':
            os.remove(existing_file)
    
    # Now begin to create new subpdfs
    inputpdf = PdfFileReader(open(f_path, "rb"))

    pAE = re.compile(settings.PDF_FILE_NAME_RE)
    new_pdf = True
    count = 1
    for i in range(inputpdf.numPages):
        if new_pdf:
            output = PdfFileWriter()
            
        page = inputpdf.getPage(i)
        output.addPage(page)

        content = convert_page_to_text(page)
            
        if split_key_word in content:
            saved_file_name = '{0}.pdf'.format(pAE.search(content).group(1))
            saved_file_path = os.path.join(pdf_folder, saved_file_name)
            with open(saved_file_path, 'wb') as f:
                output.write(f)
            new_pdf = True
            count += 1
        else:
            new_pdf = False
            

def get_excel_content():
    '''
    This function will read EXCEL FILE and
    get list of (ae_code, ae_name, email(1), email(2), email(3), ...)
    First item of returned excel_content is title names.
    Only works for xlsx and xlsm file format.
    '''

    f_path = settings.EXCEL_FILE
    
    wb = load_workbook(f_path)
    first_sheet_name = wb.get_sheet_names()[0]
    sheet = wb[first_sheet_name]

    title = []
    content = []
    first_row = True
    for row in sheet.iter_rows():
        row_value = []
        for cell in row:
            if first_row:
                title.append(cell.value)
            else:
                row_value.append(cell.value)

        if row_value:
            content.append(row_value)
        first_row = False
        
    file_map_key = settings.FILE_MAP_KEY
    name_key = settings.EXCEL_FILE_NAME_KEY
    ae_send = settings.AE_SEND
    email_keys = [v for v in title if v and v.lower().startswith('email')]

    excel_content = []
    for each_row in content:
        d = dict(zip(title, each_row))

        # check whether current row has file_map_key
        # if not, then skip
        file_map_value = d[file_map_key]
        if not file_map_key:
            continue

        # Check whether all emails are empty
        # if so, skip current row because no where to send emails.
        has_email = False
        emails = []
        for email_key in email_keys:
            email_address = d[email_key]
            emails.append(email_address)
            if email_address:
                has_email = True            
        if not has_email:
            continue

        excel_content.append([d[file_map_key], d[name_key], d[ae_send]] + emails)

    excel_content.sort(key=lambda x: x[0])
    return [[file_map_key, name_key, ae_send] + email_keys, ] + excel_content

def get_split_pdf_dict():
    '''
    This function will get dict of all pdf files in settings.PDF_FOLDER.

    Return:
        { pdf_file_name_without_extension: pdf_file_path, ... }
    '''
    pdf_dict = {}
    fpath = settings.PDF_FOLDER
    for f_full_name in os.listdir(fpath):
        file_path = os.path.join(fpath, f_full_name)
        f_name, f_extension = os.path.splitext(f_full_name)
        if os.path.isfile(file_path) and f_extension == '.pdf':
            pdf_dict[f_name] = file_path
            
    return pdf_dict

def _send_emails(subject, content, to_emails, pdf_file_list):
    email = EmailMessage(subject, content, settings.EMAIL_HOST_USER, to_emails)
    for pdf_file in pdf_file_list:
        email.attach_file(pdf_file, 'application/pdf')
    email.send()

def send_emails(subject, content, to_emails, pdf_file_list):
    
    # Read each file size and make attach file total size smaller than settings.MAX_ATTACH_SIZE
    pdfs = []
    total_file_size = 0 # Unit is bytes
    for pdf_file in pdf_file_list:
        file_size = os.path.getsize(pdf_file)

        if file_size >= settings.MAX_ATTACH_SIZE:
            raise Exception("File {} file size is too large.".format(file_size))
        
        total_file_size += file_size
        if total_file_size >= settings.MAX_ATTACH_SIZE:
            _send_emails(subject, content, to_emails, pdfs)
            total_file_size = file_size
            pdfs = [pdf_file, ]
        else:
            pdfs.append(pdf_file)

    if pdfs:
        _send_emails(subject, content, to_emails, pdfs)
